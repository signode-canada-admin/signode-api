from tabula import read_pdf, read_pdf_with_template
from openpyxl import Workbook
import os
import glob
from datetime import date
import sys
import json

def list_of_files(path, file_ending='*.pdf'):
    '''
    path = directory of files
    file_ending = type of files (default = ".pdf")
    '''
    return glob.glob(os.path.join(path, file_ending))

def enter_directory(path):
    '''
    path = directory to enter(must be r'str')
    '''
    try: 
        os.chdir(path)
    except OSError:       
        print("Entering the directory %s failed" % path)

def move_files(src, dst):
    '''
    src = path to file (.pdf)
    dst = new path to file (.pdf)
    '''
    try:
        os.replace(src, dst)
    except:
        os.rename(src, dst)

def paths(site):
    code_path = os.getcwd()
    root_path = rf"Y:\Pick Ticket Project\EDI" #store this to .env file
    pdf_files = os.path.join(root_path, f"CUSTOMERS\{site}")
    po_archive = os.path.join(pdf_files, "ARCHIVED_POs")
    sx_excel_path = os.path.join(root_path, "EXCEL_SX")
    sx_excel_archive = os.path.join(sx_excel_path, "ARCHIVED")
    return {
        "code": code_path,
        "root": root_path,
        "pdfs": pdf_files,
        "pdfs_archive": po_archive,
        "sx_excel": sx_excel_path,
        "sx_excel_archive": sx_excel_archive
    }

# def paths():
#     code_path = os.getcwd()
#     premium_plus_path = rf"Y:\Pick Ticket Project\EDI\Premium_plus" #store this to .env file
#     pdf_files_path_premium_plus = os.path.join(premium_plus_path, "PDFS")
#     po_archive_path_premium_plus = os.path.join(pdf_files_path_premium_plus, "ARCHIVED_PREMIUM_PLUS_POs")
#     sx_excel_path_premium_plus = os.path.join(premium_plus_path, "EXCEL_SX_PREMIUM_PLUS")
#     return {
#         "code": code_path,
#         "root": premium_plus_path,
#         "pdfs": pdf_files_path_premium_plus,
#         "pdfs_archive": po_archive_path_premium_plus,
#         "sx_excel": sx_excel_path_premium_plus
#     }

def get_active_sheet(excel_file_name):
    # open workbook Object
    # get active worksheet
    excel_file_extension = 'xls'
    filename = f'{date.today()}_{excel_file_name}.{excel_file_extension}'
    workbook = Workbook()
    sheet = workbook.active
    return (sheet, workbook, filename)




# print({"success": True})

# premium_plus_process({"line_items":[{"quantity":"1","product":"428756"},{"quantity":"3","product":"428789"},{"quantity":"7","product":"428811"},{"quantity":"4","product":"428873"},{"quantity":"1","product":"428895"},{"quantity":"2","product":"428909"},{"quantity":"1","product":"428920"}],"ship_via":"PUROLATOR","po_no":"411369","num_line_items":7}, '412711')

# premium_plus_process({"line_items":[{"quantity":"4","product":"2187.004"}],"ship_via":"UPS","po_no":"412749","num_line_items":1}, "412711")

# sys.stdout.flush()



def print_multiple_POS(json_data):
    '''
    json = {
     data = [{}, {}, {}, ... po_json]
    }
    excel_file_name = "test"
    
    '''
    #get active sheet and file
    (sheet, workbook, excel_file) = get_active_sheet("_".join(["_".join(name.split("*SEPARATOR*")) for name in json_data["pdfs"]]))
    # sheet.title = json_data["pdfs"][0]
    

    DATA = json_data["data"]
    
#     Reference:
    column_name = {
        "product": 1,
        "description": 2,
        "quantity": 3,
        "unit": 4,
        "price": 5,
        "discount": 6,
        "disc_type": 7,
        "vendor": 8,
        "prod_line": 9,
        "prod_cat": 10, 
        "prod_cost": 11,
        "tie_type": 12,
        "tie_whse": 13,
        "drop_ship_option": 14,
        "print_option": 15,
        "required_option": 16,
        "subtotal_option": 17,
        "print_price_option": 18,
    }
    ##print_to_header
    cell_name = {
        "customer_name": "A1",
        "ship_to": "B1",
        "warehouse": "A2",
        "order_type": "B2",
        "customer_po": "A3",
        "ship_via": "A4",
        "terms": "B4",
        "reference":"A5",
        "requested_ship_date": "B5"
        
    }
    y = 0
    
    no_pdfs = len(json_data["pdfs"])
    excel_file = excel_file.replace("_Bunzl_industrial", "Placeholder", 1)
    excel_file = excel_file.replace("_Bunzl_industrial", "", no_pdfs)
    excel_file = excel_file.replace("Placeholder", "_Bunzl_industrial")
    id = (no_pdfs) * [0] 
    for data in DATA:
        LINE_ITEMS = data["line_items"]
        MIN_ROW = 9
        MAX_ROW = MIN_ROW + data["num_line_items"] - 1
        MIN_COL = 1
        MAX_COL = 18

        sheet.title = data["po_no"]
        
        try:
            sheet[cell_name["terms"]] = data["terms"]
            sheet[cell_name["reference"]] = data["reference"]
            sheet[cell_name["customer_name"]] = data["customer_name"]
        except:
            sheet[cell_name["customer_name"]] = data["customer_name"]
        sheet[cell_name["ship_to"]] = data["ship_to"]
        sheet[cell_name["warehouse"]] = data["warehouse"]
        sheet[cell_name["order_type"]] = data["order_type"]
        sheet[cell_name["customer_po"]] = data["po_no"]
        sheet[cell_name["ship_via"]] = data["ship_via"]

        id[y] = data["_id"]
        y = y + 1
            
#         sheet[cell_name["terms"]] = "PPD"
    #     sheet[cell_name["requested_ship_date"]] = "NET 30 DAYS"

        ##print_to_line_items
        quantity_query = column_name['quantity'] - 1
        product_query = column_name['product'] -1 
        description_query = column_name['description'] - 1
        if int(data["customer_name"]) == 6207:
            for item, row in enumerate(sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL, values_only=False)):
                try:
                    row[product_query].value = LINE_ITEMS[item]['product']
                    row[quantity_query].value = LINE_ITEMS[item]['quantity']
                except:
                    pass
                try:
                    row[description_query].value = (LINE_ITEMS[item]['description'])
                except:
                    pass
                
        else:
            for item, row in enumerate(sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL, values_only=False)):
                row[product_query].value = LINE_ITEMS[item]['product']
                row[quantity_query].value = LINE_ITEMS[item]['quantity']
                try:
                    row[description_query].value = (LINE_ITEMS[item]['description'])
                except:
                    pass


        if len(workbook.sheetnames) != len(DATA):
            sheet = workbook.create_sheet()
        #save the excel sheet
    workbook.save(filename=excel_file)
    return json_data["pdfs"], id


# def move_pdf(site):
#     pdf = os.path.join(paths[site]['pdfs'], f'{filename}.pdf')
#     move_files(pdf, os.path.join(all_paths["pdfs_archive"], f'{filename}.pdf'))
# def move_PO(lst=[]):


def premium_plus_process(data):
    
    all_paths = paths("Premium_plus")
    # enter excel directory
    enter_directory(all_paths["sx_excel"])
    files, x = print_multiple_POS(data)
    # move pdf file to archive
    site = (files[0]).split("*SEPARATOR*")
    test = 'intial'
    subtitles = ['-Consumables', '-Parts', '-Other']

    
    if (files[0]).split("*SEPARATOR*")[1] == 'Service':
        for i in range(0,len(files)):
            all_paths = paths('Service')
            pdf = os.path.join(all_paths['pdfs'], f'{x[i]}.pdf')
            move_files(pdf, os.path.join(all_paths["pdfs_archive"], f'{x[i]}.pdf'))
    if site[1] == 'Bunzl_industrial':
        for filename in files:
            id , site = filename.split("*SEPARATOR*")
            for a in subtitles:
                if a in id:
                    id = (id.split(a))[0]
            print(id)
            all_paths = paths(site)
            pdf = os.path.join(all_paths['pdfs'], f'{id}.pdf')
            if test != id:
                move_files(pdf, os.path.join(all_paths["pdfs_archive"], f'{id}.pdf'))
            else:
                pass
            test = id
    else:
        for filename in files:
            id , site = filename.split("*SEPARATOR*")
            all_paths = paths(site)
            pdf = os.path.join(all_paths['pdfs'], f'{id}.pdf')
            move_files(pdf, os.path.join(all_paths["pdfs_archive"], f'{id}.pdf'))
    # finally enter code_path
    enter_directory(all_paths["code"])
    # return "DONE"
    return "DONE"

try:
    data = premium_plus_process(json.loads(sys.argv[1]))
    print({
        "success": "true",
        "data": data
    })
except Exception as e: 
    print({
        "success": "false",
        "error": "Unexpected Error occured"
    })
sys.stdout.flush()

'''
premium_plus_process({"data": [{'ship_via': 'PUROLATOR',
  'po_no': '411369',
     "customer_name": 6183,
     "ship_to": "00",
     "warehouse": "A001",
     "order_type": "QU",
  'num_line_items': 7,
  'line_items': [{'quantity': 1, 'product': '428756'},
   {'quantity': 3, 'product': '428789'},
   {'quantity': 7, 'product': '428811'},
   {'quantity': 4, 'product': '428873'},
   {'quantity': 1, 'product': '428895'},
   {'quantity': 2, 'product': '428909'},
   {'quantity': 1, 'product': '428920'}]}, 
                               {
     "ship_via": "UPS",
     "po_no": "412749",
     "customer_name": 2535,
 "ship_to": "00",
 "warehouse": "A001",
 "order_type": "QU",
     "num_line_items": 1,
     "line_items": [
       {
         "quantity": 4,
         "product": "2187.004"
       }
     ]
   }], 
   "pdfs": ["412749*SEPARATOR*Premium_plus",
 "Commande Signode-SRAP YL-813*SEPARATOR*Srap"]})
'''
